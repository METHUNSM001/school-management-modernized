import openpyxl
import os
from datetime import datetime

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'data')

SCHEMAS = {
    'users':       ['id','username','password','role','name','email','phone','status','created_at'],
    'students':    ['id','user_id','name','roll_no','class_id','section','dob','gender','address','parent_name','parent_phone','parent_email','admission_date','photo','status'],
    'teachers':    ['id','user_id','name','employee_id','subject','qualification','phone','email','address','joining_date','salary','photo','status'],
    'classes':     ['id','name','section','teacher_id','capacity','academic_year'],
    'subjects':    ['id','name','code','class_id','teacher_id','max_marks'],
    'attendance':  ['id','student_id','class_id','date','status','marked_by'],
    'assignments': ['id','title','description','class_id','subject_id','teacher_id','due_date','file_path','created_at'],
    'submissions': ['id','assignment_id','student_id','file_path','submitted_at','marks','feedback'],
    'exams':       ['id','name','class_id','subject_id','exam_date','max_marks','passing_marks','created_by'],
    'marks':       ['id','exam_id','student_id','marks_obtained','grade','remarks'],
    'fees':        ['id','student_id','fee_type','amount','due_date','status','paid_date','receipt_no','academic_year'],
    'events':      ['id','title','description','event_date','location','photo','created_by','created_at'],
    'announcements':['id','title','content','target_role','created_by','created_at','is_active'],
    'library_books':['id','title','author','isbn','category','copies','available','added_date'],
    'issued_books': ['id','book_id','student_id','issue_date','due_date','return_date','fine'],
    'transport':   ['id','route_name','vehicle_no','driver_name','driver_phone','stops','students'],
    'gallery':     ['id','title','category','file_path','uploaded_by','uploaded_at'],
    'timetable':   ['id','class_id','day','period','subject_id','teacher_id','start_time','end_time'],
    'admissions':  ['id','applicant_name','dob','gender','class_applying','parent_name','parent_phone','email','address','prev_school','status','applied_date'],
    'contact_msgs':['id','name','email','phone','subject','message','received_at','is_read'],
}

def get_wb(name):
    path = os.path.join(DATA_DIR, f'{name}.xlsx')
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = name
        ws.append(SCHEMAS[name])
        wb.save(path)
    return openpyxl.load_workbook(path), path

def read_all(name):
    wb, _ = get_wb(name)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        return []
    headers = rows[0]
    return [dict(zip(headers, r)) for r in rows[1:]]

def read_by_id(name, rid):
    all_rows = read_all(name)
    for r in all_rows:
        if str(r.get('id')) == str(rid):
            return r
    return None

def read_where(name, **kwargs):
    all_rows = read_all(name)
    result = []
    for r in all_rows:
        match = all(str(r.get(k,'')) == str(v) for k, v in kwargs.items())
        if match:
            result.append(r)
    return result

def next_id(name):
    all_rows = read_all(name)
    if not all_rows:
        return 1
    ids = [int(r['id']) for r in all_rows if r.get('id') and str(r['id']).isdigit()]
    return max(ids) + 1 if ids else 1

def insert(name, data):
    wb, path = get_wb(name)
    ws = wb.active
    headers = SCHEMAS[name]
    row = [data.get(h, '') for h in headers]
    ws.append(row)
    wb.save(path)
    return data

def update(name, rid, data):
    wb, path = get_wb(name)
    ws = wb.active
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    id_col = list(headers).index('id') + 1
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col-1].value) == str(rid):
            for key, val in data.items():
                if key in headers:
                    col = list(headers).index(key) + 1
                    row[col-1].value = val
            break
    wb.save(path)

def delete(name, rid):
    wb, path = get_wb(name)
    ws = wb.active
    headers = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    id_col = list(headers).index('id') + 1
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if str(row[id_col-1].value) == str(rid):
            ws.delete_rows(i)
            break
    wb.save(path)

def count(name):
    return len(read_all(name))

def init_data():
    """Seed initial data"""
    os.makedirs(DATA_DIR, exist_ok=True)
    # Init all sheets
    for sheet in SCHEMAS:
        get_wb(sheet)
    
    # Create default admin if no users
    users = read_all('users')
    if not users:
        from werkzeug.security import generate_password_hash
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        insert('users', {'id':1,'username':'admin','password':generate_password_hash('admin123'),
                          'role':'admin','name':'System Admin','email':'admin@school.edu',
                          'phone':'9999999999','status':'active','created_at':now})
        insert('users', {'id':2,'username':'teacher1','password':generate_password_hash('teacher123'),
                          'role':'teacher','name':'Mrs. Priya Sharma','email':'priya@school.edu',
                          'phone':'9888888888','status':'active','created_at':now})
        insert('users', {'id':3,'username':'student1','password':generate_password_hash('student123'),
                          'role':'student','name':'Arjun Kumar','email':'arjun@school.edu',
                          'phone':'9777777777','status':'active','created_at':now})
        insert('users', {'id':4,'username':'parent1','password':generate_password_hash('parent123'),
                          'role':'parent','name':'Rajesh Kumar','email':'rajesh@school.edu',
                          'phone':'9666666666','status':'active','created_at':now})
        
        # Classes
        classes_data = [
            'PRE-KG', 'LKG', 'UKG', 'Class 1', 'Class 2', 'Class 3', 'Class 4', 
            'Class 5', 'Class 6', 'Class 7', 'Class 8', 'Class 9', 'Class 10',
            'Class 11 - BIO/MATHS', 'Class 11 - BIO/CS', 'Class 11 - MATHS/CS', 'Class 11 - ARTS',
            'Class 12 - BIO/MATHS', 'Class 12 - BIO/CS', 'Class 12 - MATHS/CS', 'Class 12 - ARTS'
        ]
        for i, cname in enumerate(classes_data, 1):
            insert('classes', {'id':i,'name':cname,'section':'A','teacher_id':1,'capacity':40,'academic_year':'2024-25'})
        
        # Subjects
        insert('subjects', {'id':1,'name':'Mathematics','code':'MATH10','class_id':1,'teacher_id':1,'max_marks':100})
        insert('subjects', {'id':2,'name':'Science','code':'SCI10','class_id':1,'teacher_id':1,'max_marks':100})
        insert('subjects', {'id':3,'name':'English','code':'ENG10','class_id':1,'teacher_id':1,'max_marks':100})
        insert('subjects', {'id':4,'name':'Social Studies','code':'SS10','class_id':1,'teacher_id':1,'max_marks':100})
        insert('subjects', {'id':5,'name':'Hindi','code':'HIN10','class_id':1,'teacher_id':1,'max_marks':100})
        
        # Teacher record
        insert('teachers', {'id':1,'user_id':2,'name':'Mrs. Priya Sharma','employee_id':'TCH001',
                             'subject':'Mathematics','qualification':'M.Sc, B.Ed','phone':'9888888888',
                             'email':'priya@school.edu','address':'Chennai','joining_date':'2020-06-01',
                             'salary':45000,'photo':'','status':'active'})
        
        # Student record
        insert('students', {'id':1,'user_id':3,'name':'Arjun Kumar','roll_no':'10A001','class_id':1,
                             'section':'A','dob':'2008-05-15','gender':'Male','address':'Chennai',
                             'parent_name':'Rajesh Kumar','parent_phone':'9666666666',
                             'parent_email':'rajesh@school.edu','admission_date':'2020-06-01',
                             'photo':'','status':'active'})
        
        # Fees
        insert('fees', {'id':1,'student_id':1,'fee_type':'Tuition Fee','amount':12000,
                         'due_date':'2024-04-30','status':'paid','paid_date':'2024-04-10',
                         'receipt_no':'RCP001','academic_year':'2024-25'})
        insert('fees', {'id':2,'student_id':1,'fee_type':'Exam Fee','amount':500,
                         'due_date':'2024-05-15','status':'pending','paid_date':'',
                         'receipt_no':'','academic_year':'2024-25'})
        
        # Announcements
        insert('announcements', {'id':1,'title':'School Reopening','content':'School will reopen on June 1st for all classes.',
                                   'target_role':'all','created_by':1,'created_at':now,'is_active':1})
        insert('announcements', {'id':2,'title':'Annual Sports Day','content':'Annual Sports Day on December 15. All students must participate.',
                                   'target_role':'all','created_by':1,'created_at':now,'is_active':1})
        
        # Events
        insert('events', {'id':1,'title':'Annual Day Celebration','description':'Annual Day celebration with cultural programs and prize distribution.',
                           'event_date':'2024-12-20','location':'School Auditorium','photo':'','created_by':1,'created_at':now})
        insert('events', {'id':2,'title':'Science Exhibition','description':'Showcase your science projects and innovations.',
                           'event_date':'2024-11-15','location':'School Labs','photo':'','created_by':1,'created_at':now})
        
        # Library books
        insert('library_books', {'id':1,'title':'Mathematics NCERT Class 10','author':'NCERT','isbn':'978-81-7450-634-4',
                                   'category':'Textbook','copies':20,'available':18,'added_date':'2023-01-01'})
        insert('library_books', {'id':2,'title':'Wings of Fire','author':'APJ Abdul Kalam','isbn':'978-81-7371-523-5',
                                   'category':'Biography','copies':5,'available':4,'added_date':'2023-01-01'})
        insert('library_books', {'id':3,'title':'The Alchemist','author':'Paulo Coelho','isbn':'978-0-06-112241-5',
                                   'category':'Fiction','copies':3,'available':3,'added_date':'2023-01-01'})
        
        # Attendance sample
        insert('attendance', {'id':1,'student_id':1,'class_id':1,'date':'2024-11-01','status':'present','marked_by':1})
        insert('attendance', {'id':2,'student_id':1,'class_id':1,'date':'2024-11-04','status':'present','marked_by':1})
        insert('attendance', {'id':3,'student_id':1,'class_id':1,'date':'2024-11-05','status':'absent','marked_by':1})
        
        # Exams
        insert('exams', {'id':1,'name':'Half Yearly Exam','class_id':1,'subject_id':1,
                          'exam_date':'2024-09-15','max_marks':100,'passing_marks':35,'created_by':1})
        insert('exams', {'id':2,'name':'Half Yearly Exam','class_id':1,'subject_id':2,
                          'exam_date':'2024-09-16','max_marks':100,'passing_marks':35,'created_by':1})
        
        # Marks
        insert('marks', {'id':1,'exam_id':1,'student_id':1,'marks_obtained':85,'grade':'A','remarks':'Excellent'})
        insert('marks', {'id':2,'exam_id':2,'student_id':1,'marks_obtained':78,'grade':'B+','remarks':'Good'})
        
        # Transport
        insert('transport', {'id':1,'route_name':'Route 1 - Anna Nagar','vehicle_no':'TN01AB1234',
                              'driver_name':'Murugan','driver_phone':'9555555555',
                              'stops':'Anna Nagar, Koyambedu, Arumbakkam','students':''})
        
        print("✅ Seed data created successfully!")
